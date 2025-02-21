import subprocess
import sys

def instalar_biblioteca(biblioteca):
    subprocess.check_call([sys.executable, "-m", "pip", "install", biblioteca])

try:
    import openpyxl
except ImportError:
    instalar_biblioteca("openpyxl")
    import openpyxl
    
import re
import math

def formatar_nome_produto(nome):
    """
    Formata o nome do produto para ser usado como nome de aba no Excel.
    Remove acentuações e caracteres especiais usando regex e substitui espaços por underscores.
    """
    substituicoes = {
        'á': 'a', 'à': 'a', 'â': 'a', 'ã': 'a', 'ä': 'a',
        'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
        'í': 'i', 'ì': 'i', 'î': 'i', 'ï': 'i',
        'ó': 'o', 'ò': 'o', 'ô': 'o', 'õ': 'o', 'ö': 'o',
        'ú': 'u', 'ù': 'u', 'û': 'u', 'ü': 'u',
        'ç': 'c',
        'Á': 'A', 'À': 'A', 'Â': 'A', 'Ã': 'A', 'Ä': 'A',
        'É': 'E', 'È': 'E', 'Ê': 'E', 'Ë': 'E',
        'Í': 'I', 'Ì': 'I', 'Î': 'I', 'Ï': 'I',
        'Ó': 'O', 'Ò': 'O', 'Ô': 'O', 'Õ': 'O', 'Ö': 'O',
        'Ú': 'U', 'Ù': 'U', 'Û': 'U', 'Ü': 'U',
        'Ç': 'C'
    }

    padrao = re.compile('|'.join(substituicoes.keys()))
    nome_sem_acentos = padrao.sub(lambda x: substituicoes[x.group()], nome)
    nome_formatado = re.sub(r'[^\w]', '_', nome_sem_acentos)
    
    return ("cartao-" + nome_formatado).upper()

def string_para_float(valor):
    """
    Converte uma string com vírgula para float.
    """
    return float(valor.replace('.', '').replace(',', '.'))

def calcular_fator(preco_vista, valor_parcela):
    """
    Calcula o fator (valor da parcela / preço à vista).
    """
    return valor_parcela / preco_vista

def calcular_taxa_juros_price(preco_vista, valor_parcela, qtd_parcelas):
    """
    Calcula a taxa de juros (Price) usando o método de Newton-Raphson.
    """
    def f(taxa):
        return valor_parcela * sum([1 / ((1.0 + taxa) ** (i + 1)) for i in range(qtd_parcelas)]) - preco_vista

    def f_derivada(taxa):
        return -valor_parcela * sum([(i + 1) / ((1.0 + taxa) ** (i + 2)) for i in range(qtd_parcelas)])

    taxa = 0.1  # Chute inicial
    for _ in range(100):  # Número máximo de iterações
        taxa_nova = taxa - f(taxa) / f_derivada(taxa)
        if abs(taxa_nova - taxa) < 1e-6:  # Precisão desejada
            break
        taxa = taxa_nova
    return taxa

def calcular_cet_anual(taxa_mensal):
    """
    Calcula o CET anual a partir da taxa mensal.
    """
    return ((1 + taxa_mensal) ** 12 - 1)

def ler_arquivo(caminho_arquivo):
    """
    Lê o arquivo de texto e retorna uma lista de listas, onde cada sublista representa uma linha do arquivo.
    Usa ponto e vírgula (;) como delimitador.
    """
    with open(caminho_arquivo, 'r', encoding='utf-8') as arquivo:
        linhas = arquivo.readlines()
    return [linha.strip().split(';') for linha in linhas]

def criar_planilhas_no_mesmo_arquivo(dados, nome_arquivo_saida):
    """
    Cria todas as planilhas em um único arquivo Excel, cada uma em uma aba separada.
    """
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for linha in dados:
        produto = formatar_nome_produto(linha[0])
        preco_vista = string_para_float(linha[1])
        qtd_parcelas = int(linha[2])
        valor_parcela = string_para_float(linha[3])

        taxa_mensal = calcular_taxa_juros_price(preco_vista, valor_parcela, qtd_parcelas)
        cet_anual = calcular_cet_anual(taxa_mensal)
        fator = calcular_fator(preco_vista, valor_parcela)

        ws = wb.create_sheet(title=produto)
        ws['A1'] = produto
        ws['B1'] = "Exemplo de informação legal. CET: {cet}% e CET a.a {cet_aa}%"
        ws.append(["Parcela", "Fator", "Juros", "IOF", "Taxa CET"])
        ws.append([qtd_parcelas, round(fator,6), round(taxa_mensal,6), 0, round(cet_anual,6)])

    wb.save(nome_arquivo_saida)
    print(f"Arquivo '{nome_arquivo_saida}' criado com sucesso!")

if __name__ == "__main__":
    caminho_arquivo = "dados.csv"
    nome_arquivo_saida = "planilhas_produtos.xlsx"

    dados = ler_arquivo(caminho_arquivo)
    criar_planilhas_no_mesmo_arquivo(dados, nome_arquivo_saida)
